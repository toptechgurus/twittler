from openpyxl import Workbook, load_workbook
import datetime, random

file_path = r"C:\Users\SUNTHAR\PycharmProjects\twitter_automation\excel_data\excel_dir\User_data_excel_TTG.xlsx"
email_list = []
mobile_list = []
password_list = []
comments_list=[]

workbook = load_workbook(file_path)

sheets = workbook.sheetnames
print(workbook.sheetnames)
print(sheets[1])

# active_sheet = workbook.active
active_sheet = workbook[sheets[0]]
#Active sheet title
# print(active_sheet.title)
max_rows = str(active_sheet.max_row)
# print(max_rows)

#grabbing excel values
# datum = active_sheet[f"A2:C{max_rows}"]
datum = active_sheet["A2:C4"]

for email_id, password, mobile in datum:
    if (email_id.value or password.value or mobile.value) is None:
        continue
    else:
        email_list.append(email_id.value)
        password_list.append(password.value)
        mobile_list.append(mobile.value)

zipped_data = zip(email_list,password_list,mobile_list)

userdata_list = list(zipped_data)
random.shuffle(userdata_list)
print(userdata_list)




active_sheet = workbook[sheets[1]]
comments_data = active_sheet["A"]

for comment in comments_data:
    if comment.value is None:
        continue
    else:
        comments_list.append(comment.value)

random.shuffle(comments_list)
print(comments_list)